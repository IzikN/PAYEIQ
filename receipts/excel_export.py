from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from .models import Receipt, Session


def export_to_excel(session_id, file_path):
    session = Session.objects.get(id=session_id)
    receipts = Receipt.objects.filter(session=session, status="Done").order_by('date_of_payment')

    wb = Workbook()
    ws = wb.active
    ws.title = "Form G Schedule"

    # Header info
    ws["A1"] = "Employer Name:"
    ws["B1"] = session.employer_name
    ws["A2"] = "State IRS:"
    ws["B2"] = session.state_irs
    ws["A3"] = "Tax Year:"
    ws["B3"] = session.tax_year

    # Table header
    headers = ["S/N", "Month", "Date of Payment", "Receipt No.", "Amount (₦)"]
    ws.append([])
    ws.append(headers)

    for col in range(1, 6):
        ws.cell(row=5, column=col).font = Font(bold=True)

    # Table data
    row_num = 6
    for i, receipt in enumerate(receipts, start=1):
        ws.cell(row=row_num, column=1).value = i
        ws.cell(row=row_num, column=2).value = receipt.month
        ws.cell(row=row_num, column=3).value = receipt.date_of_payment
        ws.cell(row=row_num, column=4).value = receipt.receipt_number
        ws.cell(row=row_num, column=5).value = receipt.amount
        row_num += 1

    # Total row
    ws.cell(row=row_num, column=4).value = "TOTAL"
    ws.cell(row=row_num, column=4).font = Font(bold=True)

    ws.cell(row=row_num, column=5).value = f"=SUM(E6:E{row_num-1})"
    ws.cell(row=row_num, column=5).font = Font(bold=True)

    # Column widths
    widths = [6, 10, 18, 20, 15]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    wb.save(file_path)